from openpyxl import *
from openpyxl.drawing.image import Image
from time import *
import random

tm = localtime(time())
sheetname = str(strftime('%Y-%m-%d | ', tm) + strftime('%H%M', tm))

wb = load_workbook("ScrimList.xlsx")
basementSheet = wb["base"]
readDataSheet = wb["rawData"]



newSheet = wb.copy_worksheet(basementSheet)
newSheet.title = sheetname
writeNewSheet = wb[sheetname]


for i in range(5):
	#레드 팀 요약
	writeNewSheet.cell(16 + i, 4, readDataSheet.cell(row = (2 + i), column=2).value)
	writeNewSheet.cell(16 + i, 6, readDataSheet.cell(row = (2 + i), column=3).value)

	#블루 팀 요약
	writeNewSheet.cell(16 + i, 10, readDataSheet.cell(row = (2 + i), column=6 ).value)
	writeNewSheet.cell(16 + i, 12, readDataSheet.cell(row = (2 + i), column=7).value)

	#레드팀 픽
	image_name = readDataSheet.cell(row = (2 + i), column=7).value 
	try:
		img = Image("Charater/" + image_name + ".png")
	except:
		img = Image("Charater/NONE.png")
	img.height = 60
	img.width = 60
	writeNewSheet.add_image(img, str(chr(74 + i))+str(11))

	#블루팀 픽
	image_name = readDataSheet.cell(row = (2 + i), column=3).value 
	try:
		img = Image("Charater/" + image_name + ".png")
	except:
		img = Image("Charater/NONE.png")
	img.height = 60
	img.width = 60
	writeNewSheet.add_image(img, str(chr(68 + i))+str(11))

	#레드팀 밴
	image_name = readDataSheet.cell(row = (9 + i), column=3).value 
	try:
		img = Image("Charater/" + image_name + ".png")
	except:
		img = Image("Charater/NONE.png")
	img.height = 60
	img.width = 60
	writeNewSheet.add_image(img, str(chr(74 + i))+str(6))

	#블루팀 밴
	image_name = readDataSheet.cell(row = (9 + i), column=2).value 
	try:
		img = Image("Charater/" + image_name + ".png")
	except:
		img = Image("Charater/NONE.png")
	img.height = 60
	img.width = 60
	writeNewSheet.add_image(img, str(chr(68 + i))+str(6))

wb.save("ScrimList.xlsx")
